#! /usr/bin/env python

import openpyxl


def main():
    excel_file = '../resources/test.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Sheet1']
    rows = sheet.max_row

    themes = []
    id_name_dict = {}
    id_name_code_tup = []
    insert_scripts = []
    update_scripts = []
    index = 1
    for i in list(range(2, rows + 1)):
        movie_name = sheet.cell(row = i, column = 1).value
        movie_code = sheet.cell(row = i, column = 2).value
        if movie_name not in themes:
            themes.append(movie_name)
            id_name_dict[movie_name] = index
            index = index + 1
        id_name_code_tup.append((id_name_dict[movie_name], movie_name, movie_code))

    for movie_name, movie_id in id_name_dict.items():
        insert_scripts.append("INSERT INTO `movie` (`id`, `name`) VALUES(%d,'%s');" % (movie_id, movie_name))

    for t in id_name_code_tup:
        update_scripts.append("UPDATE `user_movie` t SET t.`movie_id` = %d WHERE t.`code` = '%s';" % (t[0], t[2]))


    with open('../resources/init_movie.sql', 'w') as f:
        f.write('\n-- 初始化表数据\n')
        f.write('\n'.join(insert_scripts) + '\n')
        f.write('\n-- 初始化user_movie表中的movie_id\n')
        f.write('\n'.join(update_scripts) + '\n')

    print("生成初始化SQL~~")


if __name__ == '__main__':
    main()
